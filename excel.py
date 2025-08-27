import os
import sys
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import Border, Side
from pathlib import Path
import shutil
from itertools import chain

# --- Configuration for Bill of Materials (BOM) ---
BOM_CONFIG = {
    "template_filename": "bom_template.xlsx",
    "doc_type_suffix": " СП",
    "revision_key": "Ревизия СП",
    "sheet1_max_rows": 28,
    "sheetn_max_rows": 31,
    "title_block_map": {
        "Разработал": "G36",
        "Проверил": "G37",
        "Нормоконтролёр": "G39",
        "Утвердил": "G40",
        "Дата": ["J36", "J37", "J39", "J40"],
        "Наименование 1": "K37",
        "Наименование 2": "K38",
    },
    "decimal_number_cell": "K33",
    "version_cell": "K39",
    "total_sheets_cell": "S37",
    "sheet1_layout": {
        "row_ranges": [range(2, 17), range(18, 26), range(27, 30), range(31, 33)],
        "column_indices": [3, 4, 6, 8, 12, 17, 18],
    },
    "sheetn_layout": {
        "row_ranges": [range(2, 17), range(18, 22), range(23, 30), range(31, 35), range(36, 37)],
        "column_indices": [3, 4, 6, 8, 12, 13, 14],
    },
    "sheetn_decimal_cell": "K37",
    "sheetn_page_num_cell": "O39",
    # No border_settings key needed as no action is taken
}

# --- Configuration for Part List ---
PART_LIST_CONFIG = {
    "template_filename": "part_list_template.xlsx",
    "doc_type_suffix": " ПЭ3",
    "revision_key": "Ревизия ПЭ3",
    "sheet1_max_rows": 29,
    "sheetn_max_rows": 32,
    "title_block_map": {
        "Разработал": "E37",
        "Проверил": "E38",
        "Нормоконтролёр": "E40",
        "Утвердил": "E41",
        "Дата": ["H37", "H38", "H40", "H41"],
        "Наименование 1": "I38",
        "Наименование 2": "I39",
    },
    "decimal_number_cell": "I34",
    "version_cell": "I40",
    "total_sheets_cell": "P38",
    "sheet1_layout": {
        "row_ranges": [range(2, 18), range(19, 27), range(28, 31), range(32, 34)],
        "column_indices": [3, 6, 10],
    },
    "sheetn_layout": {
        "row_ranges": [range(2, 18), range(19, 23), range(24, 31), range(32, 36), range(37, 38)],
        "column_indices": [3, 6, 10],
    },
    "sheetn_decimal_cell": "I38",
    "sheetn_page_num_cell": "P40",
    # Simplified post-processing step
    "border_settings": {
        "sheet1": "Q2:Q33",
        "sheetn": "Q2:Q37"
    }
}

# Find excel files in named folder
def find_excel_files(input_dir):
    excel_files = []
    for filename in os.listdir(input_dir):
        if filename.lower().endswith(('.xls', '.xlsx')):
            excel_files.append(filename)
    return excel_files

# Read altium excel file to dataset
def read_altium_excel_file(filepath):
    try:
        # Read Excel file once and get sheet names
        xls = pd.ExcelFile(filepath)
        if not {'Sheet1', 'Sheet2'}.issubset(xls.sheet_names):
            missing = {'Sheet1', 'Sheet2'} - set(xls.sheet_names)
            print(f"Missing sheets: {', '.join(missing)}")
            sys.exit(1)

        # Read both sheets using the ExcelFile object
        list1_df = pd.read_excel(xls, sheet_name='Sheet1')
        list2_df = pd.read_excel(xls, sheet_name='Sheet2')
        return list1_df, list2_df

    except Exception as e:
        print(f"Error reading file {filepath}: {e}")
        sys.exit(1)

# Read groups excel file to dataset
def read_excel_file(filepath):
    try:
        # Read Excel file once and get sheet names
        xls = pd.ExcelFile(filepath)
        if not {'Sheet1'}.issubset(xls.sheet_names):
            print("Missing Sheet1")
            sys.exit(1)

        # Read sheet using the ExcelFile object
        df = pd.read_excel(xls, sheet_name='Sheet1')
        return df

    except Exception as e:
        print(f"Error reading file {filepath}: {e}")
        sys.exit(1)

# Write dataset to excel file
def write_to_excel(result_df, filename):
    # Define output directory and filename
    output_dir = 'output'
    filename_with_ext = f"{filename}.xlsx"
    
    try:
        # Create output directory if it doesn't exist
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        # Construct full file path
        file_path = os.path.join(output_dir, filename_with_ext)
        # Write to Excel file
        result_df.to_excel(file_path, index=False, engine='openpyxl')
        print(f"Successfully wrote {result_df.shape[0]} records to {file_path}")
        return file_path
    
    except PermissionError:
        print(f"ERROR: Permission denied - cannot write to {filename_with_ext}")
        print("Please check:")
        print("1. If the file is open in another program (close it)")
        print("2. If you have write permissions in the output folder")
        sys.exit(1)
    
    except Exception as e:
        print(f"Error writing Excel file: {str(e)}")
        sys.exit(1)

# Create a standardized filename based on parameters from a DataFrame and a config dict
def create_document_filename(df_params, config):
    # Get the common values from the DataFrame
    decimal_num = _get_param_value(df_params, "Децимальный номер")
    name = _get_param_value(df_params, "Наименование 2")
    version = _get_param_value(df_params, "Версия")

    # Get the document-specific values using keys from the config
    doc_suffix = config["doc_type_suffix"]  # " ПЭ3" or " СП"
    revision_key = config["revision_key"]   # "Ревизия ПЭ3" or "Ревизия СП"
    revision_val = _get_param_value(df_params, revision_key)

    # Assemble the filename
    filename = f"{decimal_num}{doc_suffix} ({name}) v.{version}.{revision_val}.xlsx"
    
    return filename

# Copy and rename templates
def copy_rename_template(filename, config):
    template_name = config["template_filename"]
    input_path = Path('templates') / template_name
    output_path = Path('output') / f"{filename}"
    print(f"Copying '{input_path}' to '{output_path}'")
    try:
        shutil.copy2(input_path, output_path)
        print("File copied and renamed successfully!")
    except FileNotFoundError:
        print(f"Error: The source file '{input_path}' or directory not found.")
    except PermissionError:
        print(f"Error: Permission denied to access '{input_path}' or write to '{output_path}'.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Write to merged cell
def write_to_merged_cell(ws, cell_ref, value):
    # Writes to the top-left cell of a merged range
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            # Get top-left cell of merged range
            min_col, min_row, _, _ = range_boundaries(str(merged_range))
            top_left_cell = ws.cell(row=min_row, column=min_col)
            top_left_cell.value = value
            return
    
    # If not in merged range, write normally
    ws[cell_ref] = value

def set_border_thickness(ws, range):
    # Create bold left border
    bold_left = Side(style='medium')

    # Apply to the entire range Q2:Q33
    for row in ws[range]:
        for cell in row:
            # Preserve existing borders and only modify left border
            current_border = cell.border
            cell.border = Border(
                left=bold_left,
                top=current_border.top,
                right=current_border.right,
                bottom=current_border.bottom
            )

# Get a value from the parameters DataFrame
def _get_param_value(df_params, key):
    try:
        return df_params.loc[df_params["Key"] == key, "Value"].values[0]
    except IndexError:
        raise ValueError(f"Parameter key '{key}' not found in df_params.")

# Write a chunk of data to a sheet based on the layout config
def _write_data_chunk(ws, df_data, start_row_idx, layout_config):
    row_idx = start_row_idx
    num_rows = df_data.shape[0]
    
    for excel_row in chain(*layout_config["row_ranges"]):
        if row_idx >= num_rows:
            break
        col_idx = 0
        for excel_col in layout_config["column_indices"]:
            cell_ref = f"{chr(64 + excel_col)}{excel_row}"
            cell_value = df_data.iloc[row_idx, col_idx]
            write_to_merged_cell(ws, cell_ref, cell_value)
            col_idx += 1
        row_idx += 1
    return row_idx

# Write part list or bom to template according to config dictionary
def write_document_to_template(df_params, df_data, filename, config):
    # 1. Calculate number of sheets from config
    num_rows = df_data.shape[0]
    if num_rows <= config["sheet1_max_rows"]:
        num_sheets = 1
    else:
        num_sheets = math.ceil((num_rows - config["sheet1_max_rows"]) / config["sheetn_max_rows"]) + 1

    output_path = Path('output') / filename
    print(f"Processing file: {output_path} with {num_sheets} sheet(s).")

    try:
        wb = load_workbook(output_path)
        
        # 2. Prepare multi-sheet template if needed
        if num_sheets > 1:
            if 'Sheet2' not in wb.sheetnames:
                raise ValueError("Template must contain a 'Sheet2' for multi-page output.")
            if num_sheets > 2:
                template_source = wb.copy_worksheet(wb['Sheet2'])
                template_source.title = "template_temp"

        # --- Process Sheet 1 ---
        ws = wb['Sheet1']

        # 3. Write parameters to Sheet1 title block
        for key, cell in config["title_block_map"].items():
            value = _get_param_value(df_params, key)
            if isinstance(cell, list):
                for c in cell:
                    write_to_merged_cell(ws, c, value)
            else:
                write_to_merged_cell(ws, cell, value)

        version_val = 'Версия ' + _get_param_value(df_params, "Версия") + '.' + _get_param_value(df_params, config["revision_key"])
        write_to_merged_cell(ws, config["version_cell"], version_val)
        decimal_val = _get_param_value(df_params, "Децимальный номер") + config["doc_type_suffix"]
        write_to_merged_cell(ws, config["decimal_number_cell"], decimal_val)
        write_to_merged_cell(ws, config["total_sheets_cell"], num_sheets)

        # 4. Write data to Sheet1
        current_row_idx = _write_data_chunk(ws, df_data, 0, config["sheet1_layout"])

        # 5. [MODIFIED] Run post-processing for Sheet1 if defined
        border_settings = config.get("border_settings")
        if border_settings and "sheet1" in border_settings:
            set_border_thickness(ws, border_settings["sheet1"])

        # --- Process Subsequent Sheets (SheetN) ---
        if num_sheets > 1:
            for i in range(2, num_sheets + 1):
                target_sheet_name = f"Sheet{i}"
                if i == 2:
                    ws_n = wb[target_sheet_name]
                else:
                    ws_n = wb.copy_worksheet(wb["template_temp"])
                    ws_n.title = target_sheet_name
                
                write_to_merged_cell(ws_n, config["sheetn_decimal_cell"], decimal_val)
                write_to_merged_cell(ws_n, config["sheetn_page_num_cell"], i)
                
                current_row_idx = _write_data_chunk(ws_n, df_data, current_row_idx, config["sheetn_layout"])

                if border_settings and "sheetn" in border_settings:
                    set_border_thickness(ws_n, border_settings["sheetn"])
        
        # 6. Cleanup
        if "template_temp" in wb.sheetnames:
            del wb["template_temp"]
        if num_sheets == 1 and "Sheet2" in wb.sheetnames:
            del wb["Sheet2"]
            print("Sheet 'Sheet2' deleted successfully.")

        # 7. Save Workbook
        wb.save(output_path)
        print("Workbook saved successfully.")

    except FileNotFoundError:
        print(f"Error: File or directory not found at {output_path}")
    except PermissionError:
        print(f"Error: Permission denied to write to {output_path}")
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")